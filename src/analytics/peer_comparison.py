import sqlite3
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font


DB_PATH = "db/nifty100.db"
PEER_GROUPS_PATH = "data/raw/peer_groups.xlsx"
OUTPUT_PATH = "output/peer_comparison.xlsx"


METRICS = [
    "roe",
    "roce",
    "net_profit_margin_pct",
    "debt_to_equity",
    "free_cash_flow_cr",
    "pat_cagr_5yr",
    "revenue_cagr_5yr",
    "eps_cagr_5yr",
    "interest_coverage",
    "asset_turnover",
]


def load_data():
    conn = sqlite3.connect(DB_PATH)

    ratios_query = """
        SELECT
            company_id,
            year,
            roe,
            roce,
            net_profit_margin_pct,
            debt_to_equity,
            free_cash_flow_cr,
            pat_cagr_5yr,
            revenue_cagr_5yr,
            eps_cagr_5yr,
            interest_coverage,
            asset_turnover,
            composite_quality_score
        FROM financial_ratios
        WHERE year = 'TTM'
    """

    ratios = pd.read_sql_query(ratios_query, conn)

    companies_query = """
        SELECT
            id AS company_id,
            company_name
        FROM companies
    """

    companies = pd.read_sql_query(companies_query, conn)

    conn.close()

    peers = pd.read_excel(PEER_GROUPS_PATH)

    peers.columns = [
        str(column).strip()
        for column in peers.columns
    ]

    peers["company_id"] = peers["company_id"].astype(str).str.strip()
    peers["peer_group_name"] = (
        peers["peer_group_name"]
        .astype(str)
        .str.strip()
    )

    return peers, ratios, companies


def calculate_percentile(series, inverse=False):
    values = pd.to_numeric(
        series,
        errors="coerce",
    )

    if inverse:
        return 1 - values.rank(
            pct=True,
            method="average",
        )

    return values.rank(
        pct=True,
        method="average",
    )


def build_peer_data(peers, ratios, companies):
    merged = peers.merge(
        companies,
        on="company_id",
        how="left",
    )

    merged = merged.merge(
        ratios,
        on="company_id",
        how="left",
    )

    result = []

    for peer_group in peers["peer_group_name"].unique():

        group = merged[
            merged["peer_group_name"] == peer_group
        ].copy()

        for metric in METRICS:

            if metric not in group.columns:
                group[metric] = None

            inverse = metric == "debt_to_equity"

            group[f"{metric}_percentile"] = calculate_percentile(
                group[metric],
                inverse=inverse,
            )

        result.append(group)

    if not result:
        return pd.DataFrame()

    return pd.concat(
        result,
        ignore_index=True,
    )


def format_excel(path):
    workbook = load_workbook(path)

    green_fill = PatternFill(
        fill_type="solid",
        fgColor="C6EFCE",
    )

    yellow_fill = PatternFill(
        fill_type="solid",
        fgColor="FFEB9C",
    )

    red_fill = PatternFill(
        fill_type="solid",
        fgColor="FFC7CE",
    )

    benchmark_fill = PatternFill(
        fill_type="solid",
        fgColor="FFD966",
    )

    for worksheet in workbook.worksheets:

        headers = [
            cell.value
            for cell in worksheet[1]
        ]

        percentile_columns = []

        for index, header in enumerate(headers, start=1):

            if (
                header is not None
                and str(header).endswith("_percentile")
            ):
                percentile_columns.append(index)

        for row in range(2, worksheet.max_row + 1):

            for column in percentile_columns:

                cell = worksheet.cell(
                    row=row,
                    column=column,
                )

                if cell.value is None:
                    continue

                try:
                    value = float(cell.value)

                    if value >= 0.75:
                        cell.fill = green_fill

                    elif value <= 0.25:
                        cell.fill = red_fill

                    else:
                        cell.fill = yellow_fill

                except (
                    ValueError,
                    TypeError,
                ):
                    pass

        for cell in worksheet[1]:
            cell.font = Font(
                bold=True,
            )

        for column in worksheet.columns:

            max_length = 0

            column_letter = column[0].column_letter

            for cell in column:

                if cell.value is not None:

                    max_length = max(
                        max_length,
                        len(str(cell.value)),
                    )

            worksheet.column_dimensions[
                column_letter
            ].width = min(
                max_length + 2,
                30,
            )

    workbook.save(path)


def main():

    print("=" * 60)
    print("PEER COMPARISON EXCEL REPORT")
    print("=" * 60)

    Path("output").mkdir(
        parents=True,
        exist_ok=True,
    )

    peers, ratios, companies = load_data()

    data = build_peer_data(
        peers,
        ratios,
        companies,
    )

    if data.empty:
        print("No peer comparison data found.")
        return

    with pd.ExcelWriter(
        OUTPUT_PATH,
        engine="openpyxl",
    ) as writer:

        for peer_group in peers[
            "peer_group_name"
        ].unique():

            group = data[
                data["peer_group_name"]
                == peer_group
            ].copy()

            output_columns = [
                "company_id",
                "company_name",
            ]

            for metric in METRICS:

                output_columns.append(metric)

                output_columns.append(
                    f"{metric}_percentile"
                )

            if "composite_quality_score" in group.columns:
                output_columns.append(
                    "composite_quality_score"
                )

            output_columns = [
                column
                for column in output_columns
                if column in group.columns
            ]

            group[
                output_columns
            ].to_excel(
                writer,
                sheet_name=peer_group[:31],
                index=False,
            )

    format_excel(
        OUTPUT_PATH
    )

    print()
    print("=" * 60)
    print("PEER COMPARISON COMPLETE")
    print("=" * 60)
    print(
        f"Peer groups: {peers['peer_group_name'].nunique()}"
    )
    print(
        f"Companies: {peers['company_id'].nunique()}"
    )
    print(
        f"Output file: {OUTPUT_PATH}"
    )


if __name__ == "__main__":
    main()